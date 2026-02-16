"""
Create a sample vehicle pricelist Excel file for testing.
Run: python tests/create_test_excel.py
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.styles import Font, Alignment

def create_sample_pricelist():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Toyota Sedan Range ---
    ws = wb.active
    ws.title = "Sedan Range"

    # Title row (merged)
    ws.merge_cells("A1:F1")
    ws["A1"] = "Toyota Price List 2026 – Sedan Range"
    ws["A1"].font = Font(bold=True, size=14)

    # Headers
    headers = ["Model", "Engine", "Transmission", "Drivetrain", "Price (EUR)", "Doors"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)

    # Data rows
    rows = [
        ("Corolla", "1.8L Hybrid", "CVT", "FWD", 28950, 4),
        ("Corolla", "2.0L Hybrid", "CVT", "FWD", 32450, 4),
        ("Camry", "2.5L Hybrid", "CVT", "FWD", 39900, 4),
        ("Camry", "2.5L Hybrid AWD", "CVT", "AWD", 42500, 4),
    ]
    for i, row_data in enumerate(rows, 4):
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)

    # Footnote
    ws.merge_cells("A9:F9")
    ws["A9"] = "* Prices exclude VAT. Valid from 01-Jan-2026."
    ws["A9"].font = Font(italic=True, size=9)

    # --- Sheet 2: SUV Range ---
    ws2 = wb.create_sheet("SUV Range")

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "Toyota Price List 2026 – SUV Range"
    ws2["A1"].font = Font(bold=True, size=14)

    headers2 = ["Model", "Variant", "Engine", "Power (HP)", "Transmission", "Price (EUR)", "Seats"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)

    suv_rows = [
        ("RAV4", "Active", "2.5L Hybrid", 222, "CVT", 38500, 5),
        ("RAV4", "Style", "2.5L Hybrid", 222, "CVT", 42000, 5),
        ("RAV4", "Style AWD", "2.5L Plug-in Hybrid", 306, "CVT", 51900, 5),
        ("Highlander", "Comfort", "2.5L Hybrid", 248, "CVT", 52900, 7),
        ("Highlander", "Executive", "2.5L Hybrid", 248, "CVT", 59900, 7),
        ("Land Cruiser", "Active", "2.8L Diesel", 204, "Auto 8-speed", 69900, 7),
    ]
    for i, row_data in enumerate(suv_rows, 4):
        for col, val in enumerate(row_data, 1):
            ws2.cell(row=i, column=col, value=val)

    # --- Sheet 3: German market (multilingual test) ---
    ws3 = wb.create_sheet("Preisliste DE")

    ws3.merge_cells("A1:F1")
    ws3["A1"] = "Toyota Preisliste 2026 – Deutschland"
    ws3["A1"].font = Font(bold=True, size=14)

    headers3 = ["Modell", "Motor", "Getriebe", "Antrieb", "Preis (EUR)", "Türen"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)

    de_rows = [
        ("Corolla", "1.8L Hybrid", "CVT", "Frontantrieb", 29450, 4),
        ("Corolla", "2.0L Hybrid", "CVT", "Frontantrieb", 33150, 4),
        ("Yaris", "1.5L Hybrid", "CVT", "Frontantrieb", 22900, 5),
    ]
    for i, row_data in enumerate(de_rows, 4):
        for col, val in enumerate(row_data, 1):
            ws3.cell(row=i, column=col, value=val)

    # Save
    out_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "tests", "fixtures")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "toyota_pricelist_2026.xlsx")
    wb.save(out_path)
    print(f"Created: {out_path}")
    return out_path


if __name__ == "__main__":
    create_sample_pricelist()
