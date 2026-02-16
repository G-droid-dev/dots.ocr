"""
Excel → HTML table conversion using openpyxl.
Handles merged cells, multi-sheet workbooks, and common formatting.
"""

import os
from typing import List, Tuple

import openpyxl
from openpyxl.utils import get_column_letter


def excel_to_html_tables(file_path: str) -> List[Tuple[str, str]]:
    """
    Convert an Excel workbook into a list of (sheet_name, html_table_string) tuples.

    Each sheet becomes one HTML ``<table>`` element.  Merged cells are handled
    by emitting the value only in the top-left cell and using ``colspan`` /
    ``rowspan`` attributes; every other cell covered by the merge is skipped.

    Parameters
    ----------
    file_path : str
        Absolute path to an .xlsx or .xls file.

    Returns
    -------
    list of (str, str)
        Each element is ``(sheet_name, html_string)``.
    """
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    results: List[Tuple[str, str]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        html = _sheet_to_html(ws)
        if html:  # skip completely empty sheets
            results.append((sheet_name, html))

    wb.close()
    return results


def _sheet_to_html(ws) -> str:
    """Convert a single openpyxl worksheet to an HTML table string."""
    if ws.max_row is None or ws.max_column is None:
        return ""
    if ws.max_row == 0 or ws.max_column == 0:
        return ""

    # Build a set of cells that are "covered" by a merge (i.e. not top-left)
    merge_map = {}  # (row, col) -> (rowspan, colspan)  for top-left cells
    skip_cells = set()  # cells to skip because they are inside a merge

    for merge_range in ws.merged_cells.ranges:
        min_row, min_col = merge_range.min_row, merge_range.min_col
        max_row, max_col = merge_range.max_row, merge_range.max_col
        rowspan = max_row - min_row + 1
        colspan = max_col - min_col + 1
        merge_map[(min_row, min_col)] = (rowspan, colspan)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) != (min_row, min_col):
                    skip_cells.add((r, c))

    rows_html = []
    has_content = False

    for row_idx in range(1, ws.max_row + 1):
        cells_html = []
        row_has_content = False

        for col_idx in range(1, ws.max_column + 1):
            if (row_idx, col_idx) in skip_cells:
                continue

            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None:
                value = ""
            else:
                value = str(value).strip()
                if value:
                    has_content = True
                    row_has_content = True

            # Use <th> for the first row (header heuristic)
            tag = "th" if row_idx == 1 else "td"

            attrs = ""
            if (row_idx, col_idx) in merge_map:
                rspan, cspan = merge_map[(row_idx, col_idx)]
                if rspan > 1:
                    attrs += f' rowspan="{rspan}"'
                if cspan > 1:
                    attrs += f' colspan="{cspan}"'

            # Escape basic HTML entities
            value = _html_escape(value)
            cells_html.append(f"<{tag}{attrs}>{value}</{tag}>")

        if cells_html:
            rows_html.append("<tr>" + "".join(cells_html) + "</tr>")

    if not has_content:
        return ""

    return "<table>\n" + "\n".join(rows_html) + "\n</table>"


def _html_escape(text: str) -> str:
    """Minimal HTML entity escaping."""
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )
